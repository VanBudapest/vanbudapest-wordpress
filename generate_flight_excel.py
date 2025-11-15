#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Járatlista Excel generáló script
Létrehoz egy Excel fájlt a megadott járatlistával és ellenőrzési megjegyzésekkel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Járatok adatai (Repülőtér, Dátum, Időpont, Járatszám)
flights_data = [
    ("BUD", "22.11.2025", "8:15:00", "TK1035", "Helyes - Istanbul->Budapest, érkezés ~08:00-08:20"),
    ("BUD", "22.11.2025", "11:15:00", "KL1363", "Időpont eltérés - Valós érkezés: 10:50 (25 perc eltérés)"),
    ("BUD", "22.11.2025", "11:35:00", "LH 1336", "HIBÁS IDŐPONT! Valós érkezés: 13:35 (2 óra eltérés)"),
    ("BUD", "22.11.2025", "13:50:00", "LH1338", "HIBÁS IDŐPONT! Valós érkezés: 17:50 (4 óra eltérés)"),
    ("BUD", "22.11.2025", "16:10:00", "BA 868", "HIBÁS IDŐPONT! Valós érkezés: 13:20 (2h50p korábban)"),
    ("BUD", "22.11.2025", "16:15:00", "LH1678", "Helyes - Munich->Budapest, érkezés 16:15"),
    ("BUD", "22.11.2025", "17:10:00", "LO 537", "Időpont eltérés - Valós érkezés: 16:55 (15 perc eltérés)"),
    ("BUD", "22.11.2025", "19:10:00", "LX2258", "Helyes - Zurich->Budapest, érkezés ~19:10-19:15"),
    ("BUD", "22.11.2025", "21:55:00", "OS639", "❌ HIBÁS JÁRAT! NEM BUDAPESTRE JÖN! Vienna->Tbilisi járat"),
    ("BUD", "22.11.2025", "22:00:00", "BA 872", "Helyes - London->Budapest, érkezés ~22:00"),
    ("BUD", "22.11.2025", "22:00:00", "BA 872", "DUPLIKÁLT! Ugyanez a járat már szerepel fent"),
    ("BUD", "22.11.2025", "23:10:00", "KL1371", "Időpont eltérés - Valós érkezés: 23:05 (5 perc eltérés)")
]

# Excel munkafüzet létrehozása
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Járatlista BUD 2025.11.22"

# Fejléc stílus
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')

# Szegély stílus
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Fejléc beállítása
headers = ['Repülőtér', 'Dátum', 'Időpont', 'Járatszám', 'Megjegyzés']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Adatok beillesztése
for row_num, flight in enumerate(flights_data, 2):
    for col_num, value in enumerate(flight, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = thin_border

        # Középre igazítás az első 4 oszlopban
        if col_num <= 4:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Hibás sorok kiemelése pirossal
        if "HIBÁS" in str(flight[4]) or "❌" in str(flight[4]):
            cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        # Duplikált sorok kiemelése sárgával
        elif "DUPLIKÁLT" in str(flight[4]):
            cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        # Időpont eltérések narancssárgával
        elif "eltérés" in str(flight[4]) and "HIBÁS" not in str(flight[4]):
            cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
        # Helyes sorok zölddel
        elif "Helyes" in str(flight[4]):
            cell.fill = PatternFill(start_color='E6F4EA', end_color='E6F4EA', fill_type='solid')

# Oszlopszélességek beállítása
ws.column_dimensions['A'].width = 12  # Repülőtér
ws.column_dimensions['B'].width = 13  # Dátum
ws.column_dimensions['C'].width = 11  # Időpont
ws.column_dimensions['D'].width = 12  # Járatszám
ws.column_dimensions['E'].width = 70  # Megjegyzés

# Sorok magasságának automatikus beállítása
for row in ws.iter_rows(min_row=2, max_row=len(flights_data) + 1):
    ws.row_dimensions[row[0].row].height = 30

ws.row_dimensions[1].height = 25  # Fejléc magasság

# Fájl mentése
output_file = "/home/user/vanbudapest-wordpress/Jaratlista_BUD_20251122.xlsx"
wb.save(output_file)
print(f"Excel fájl sikeresen létrehozva: {output_file}")
